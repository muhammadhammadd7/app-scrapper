"""
╔══════════════════════════════════════════════════════════════════╗
║     WORDPRESS & SHOPIFY WEBSITE SCRAPER — ALL 50 CATEGORIES     ║
║  • Uses DuckDuckGo HTML (no API key needed)                     ║
║  • Verifies each site is actually WordPress or Shopify          ║
║  • 5 WordPress + 5 Shopify per category                         ║
║  • Saves progress every 5 categories (auto-resume)              ║
║  • Saves Excel + Word Doc at the end                            ║
╠══════════════════════════════════════════════════════════════════╣
║  SETUP (run once):                                              ║
║    python3 -m pip install requests openpyxl                     ║
║                                                                 ║
║  RUN:                                                           ║
║    python3 wp_shopify_all_categories.py                         ║
╚══════════════════════════════════════════════════════════════════╝
"""

import time, json, re, warnings, urllib.parse, subprocess, os
warnings.filterwarnings("ignore")

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f"\n❌ Missing: {e}")
    print("Run: python3 -m pip install requests openpyxl")
    exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
LINKS_PER_CAT  = 5
VERIFY_TIMEOUT = 8
SEARCH_DELAY   = 2.0
VERIFY_DELAY   = 0.6
PROGRESS_FILE  = "_wp_shopify_progress.json"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

SKIP = [
    "google.","youtube.","facebook.","twitter.","linkedin.","instagram.",
    "wikipedia.","amazon.","reddit.","pinterest.","yelp.","tripadvisor.",
    "play.google.","apps.apple.","web.archive.","ebay.","etsy.","w3schools.",
    "stackoverflow.","github.","medium.","quora.","msn.","apple.com",
    "bing.com","duckduckgo.","yahoo.","wp.com","wordpress.com/support",
]

# ── 50 Categories with targeted queries ──────────────────────────────────────
CATEGORIES = {
    "Accounting or Book Keeping Website": {
        "wp": ["bookkeeping firm wordpress website -site:wordpress.com", "accounting small business wordpress inurl:wp-content"],
        "sh": ["bookkeeping accounting store site:myshopify.com", "accounting tools shopify store"],
    },
    "Agriculture or Farming Website": {
        "wp": ["organic farm wordpress website inurl:wp-content", "agriculture farming blog wordpress site"],
        "sh": ["farm fresh produce store site:myshopify.com", "organic farming shopify shop"],
    },
    "ALL WEBSITE": {
        "wp": ["multipurpose business wordpress website inurl:wp-content", "portfolio agency wordpress site"],
        "sh": ["general store site:myshopify.com products", "multi category shopify store"],
    },
    "Art or Sketch Website": {
        "wp": ["artist portfolio wordpress website inurl:wp-content", "digital art gallery wordpress site"],
        "sh": ["art prints shop site:myshopify.com", "illustration art store shopify"],
    },
    "Baby Website": {
        "wp": ["baby products blog wordpress inurl:wp-content", "newborn parenting wordpress website"],
        "sh": ["baby products store site:myshopify.com", "newborn baby shop shopify"],
    },
    "Bakery and Cake Website": {
        "wp": ["cake bakery wordpress website inurl:wp-content", "custom cake shop wordpress site"],
        "sh": ["custom cake shop site:myshopify.com", "bakery cake shopify store"],
    },
    "Bakery Website": {
        "wp": ["artisan bakery wordpress website inurl:wp-content", "bread bakery shop wordpress site"],
        "sh": ["bakery bread shop site:myshopify.com", "artisan bread shopify store"],
    },
    "Bicycle Websites": {
        "wp": ["bicycle shop wordpress website inurl:wp-content", "cycling blog wordpress site"],
        "sh": ["bicycle shop site:myshopify.com", "cycling gear store shopify"],
    },
    "Bloging Website": {
        "wp": ["personal blog wordpress website inurl:wp-content", "lifestyle blog wordpress site"],
        "sh": ["blogger merchandise store site:myshopify.com", "content creator shopify shop"],
    },
    "Business or Investment Website": {
        "wp": ["business consulting wordpress website inurl:wp-content", "investment firm wordpress site"],
        "sh": ["business tools store site:myshopify.com", "investment products shopify"],
    },
    "Candy Website": {
        "wp": ["candy sweet shop wordpress website inurl:wp-content", "confectionery blog wordpress site"],
        "sh": ["candy sweet shop site:myshopify.com", "chocolate candy store shopify"],
    },
    "Car or Automobile Website": {
        "wp": ["auto parts shop wordpress website inurl:wp-content", "car dealership wordpress site"],
        "sh": ["car parts accessories site:myshopify.com", "auto accessories shopify store"],
    },
    "Care Website": {
        "wp": ["elderly home care wordpress website inurl:wp-content", "caregiving services wordpress site"],
        "sh": ["elder care products site:myshopify.com", "senior care store shopify"],
    },
    "CBD Oil Website": {
        "wp": ["cbd oil hemp wordpress website inurl:wp-content", "cbd wellness blog wordpress site"],
        "sh": ["cbd oil hemp store site:myshopify.com", "cbd wellness shopify store"],
    },
    "Charity Foundation Website": {
        "wp": ["nonprofit charity wordpress website inurl:wp-content", "charity foundation wordpress site"],
        "sh": ["charity donation store site:myshopify.com", "nonprofit shop shopify"],
    },
    "Cleaning website": {
        "wp": ["cleaning services company wordpress website inurl:wp-content", "house cleaning wordpress site"],
        "sh": ["cleaning supplies store site:myshopify.com", "eco cleaning products shopify"],
    },
    "Clothing Website": {
        "wp": ["clothing fashion brand wordpress website inurl:wp-content", "apparel store wordpress site"],
        "sh": ["clothing fashion store site:myshopify.com", "apparel brand shopify"],
    },
    "Coaching Website": {
        "wp": ["life coach wordpress website inurl:wp-content", "business coaching wordpress site"],
        "sh": ["coaching courses store site:myshopify.com", "life coaching products shopify"],
    },
    "Consultant Website": {
        "wp": ["business consultant wordpress website inurl:wp-content", "consulting agency wordpress site"],
        "sh": ["consulting services store site:myshopify.com", "consultant tools shopify"],
    },
    "Cosmetic or Beauty Website": {
        "wp": ["beauty cosmetics brand wordpress website inurl:wp-content", "makeup skincare wordpress site"],
        "sh": ["cosmetics beauty store site:myshopify.com", "makeup beauty brand shopify"],
    },
    "Costume or Swimwear Website": {
        "wp": ["costume swimwear shop wordpress website inurl:wp-content", "swimwear brand wordpress site"],
        "sh": ["swimwear costume store site:myshopify.com", "costume shop shopify"],
    },
    "Dating Websites": {
        "wp": ["dating service wordpress website inurl:wp-content", "matchmaking dating wordpress site"],
        "sh": ["romantic gifts store site:myshopify.com", "couples gifts shopify"],
    },
    "Digital or Marketing Website": {
        "wp": ["digital marketing agency wordpress website inurl:wp-content", "seo marketing firm wordpress site"],
        "sh": ["marketing tools store site:myshopify.com", "digital marketing products shopify"],
    },
    "Doctor or Clinic Website": {
        "wp": ["medical clinic wordpress website inurl:wp-content", "doctor healthcare wordpress site"],
        "sh": ["medical health products site:myshopify.com", "clinic health store shopify"],
    },
    "Dog Websites": {
        "wp": ["dog pet care wordpress website inurl:wp-content", "dog training blog wordpress site"],
        "sh": ["dog pet store site:myshopify.com", "dog accessories shopify"],
    },
    "Ecommerce Website": {
        "wp": ["woocommerce online store wordpress website inurl:wp-content", "ecommerce shop wordpress site"],
        "sh": ["online store site:myshopify.com products", "ecommerce store shopify"],
    },
    "Education Website": {
        "wp": ["online education courses wordpress website inurl:wp-content", "e-learning school wordpress site"],
        "sh": ["educational products store site:myshopify.com", "online courses shopify"],
    },
    "Engineering Website": {
        "wp": ["engineering firm wordpress website inurl:wp-content", "technical engineering blog wordpress"],
        "sh": ["engineering tools store site:myshopify.com", "technical equipment shopify"],
    },
    "Event Planning Website": {
        "wp": ["event planning company wordpress website inurl:wp-content", "wedding event planner wordpress site"],
        "sh": ["event party supplies site:myshopify.com", "wedding supplies shopify store"],
    },
    "Fitness-Gym Website": {
        "wp": ["fitness gym wordpress website inurl:wp-content", "personal trainer wordpress site"],
        "sh": ["fitness gym equipment site:myshopify.com", "workout gear shopify store"],
    },
    "Food-Resturant Website": {
        "wp": ["restaurant food wordpress website inurl:wp-content", "food catering wordpress site"],
        "sh": ["gourmet food store site:myshopify.com", "food delivery shopify store"],
    },
    "Furniture Websites": {
        "wp": ["furniture home decor wordpress website inurl:wp-content", "interior furniture wordpress site"],
        "sh": ["furniture home decor site:myshopify.com", "modern furniture shopify store"],
    },
    "Gaming Website": {
        "wp": ["gaming gear wordpress website inurl:wp-content", "video game blog wordpress site"],
        "sh": ["gaming merchandise site:myshopify.com", "gaming gear store shopify"],
    },
    "Health or Nutrition Website": {
        "wp": ["health nutrition blog wordpress website inurl:wp-content", "wellness health wordpress site"],
        "sh": ["health supplements store site:myshopify.com", "nutrition wellness shopify"],
    },
    "Hotel or Traveling Website": {
        "wp": ["travel blog wordpress website inurl:wp-content", "hotel travel agency wordpress site"],
        "sh": ["travel accessories store site:myshopify.com", "luggage travel shopify store"],
    },
    "Jewellery Website": {
        "wp": ["jewellery jewelry wordpress website inurl:wp-content", "handmade jewelry wordpress site"],
        "sh": ["jewellery jewelry store site:myshopify.com", "handmade jewelry shopify"],
    },
    "Logistics Website": {
        "wp": ["logistics delivery company wordpress website inurl:wp-content", "freight logistics wordpress site"],
        "sh": ["shipping logistics store site:myshopify.com", "logistics tools shopify"],
    },
    "Music or Event Website": {
        "wp": ["music band wordpress website inurl:wp-content", "event concert wordpress site"],
        "sh": ["music merchandise store site:myshopify.com", "band merch shopify store"],
    },
    "Perfume Website": {
        "wp": ["perfume fragrance wordpress website inurl:wp-content", "fragrance brand blog wordpress site"],
        "sh": ["perfume fragrance store site:myshopify.com", "fragrance brand shopify"],
    },
    "Photography Website": {
        "wp": ["photography studio wordpress website inurl:wp-content", "photographer portfolio wordpress site"],
        "sh": ["photography prints store site:myshopify.com", "camera gear shopify store"],
    },
    "Plumbing or Heating Website": {
        "wp": ["plumbing heating company wordpress website inurl:wp-content", "hvac plumbing wordpress site"],
        "sh": ["plumbing supplies store site:myshopify.com", "heating plumbing shopify"],
    },
    "Real Estate or Construction Website": {
        "wp": ["real estate construction wordpress website inurl:wp-content", "property real estate wordpress site"],
        "sh": ["construction materials store site:myshopify.com", "real estate products shopify"],
    },
    "Recruitment or Job Finding Website": {
        "wp": ["recruitment job board wordpress website inurl:wp-content", "career job finding wordpress site"],
        "sh": ["career tools store site:myshopify.com", "resume tools shopify store"],
    },
    "Saloon Spa Website": {
        "wp": ["salon spa wordpress website inurl:wp-content", "hair beauty salon wordpress site"],
        "sh": ["salon spa products site:myshopify.com", "hair salon products shopify"],
    },
    "SKIN OR HAIR WEBSITE": {
        "wp": ["skincare hair care wordpress website inurl:wp-content", "hair skin beauty wordpress site"],
        "sh": ["skincare hair store site:myshopify.com", "hair care products shopify"],
    },
    "Social Media Website": {
        "wp": ["social media marketing wordpress website inurl:wp-content", "social media agency wordpress site"],
        "sh": ["social media tools store site:myshopify.com", "social media products shopify"],
    },
    "Solar Panel Website": {
        "wp": ["solar panel energy wordpress website inurl:wp-content", "solar renewable energy wordpress site"],
        "sh": ["solar panel products site:myshopify.com", "solar energy store shopify"],
    },
    "Sports Website": {
        "wp": ["sports club wordpress website inurl:wp-content", "sports gear blog wordpress site"],
        "sh": ["sports equipment store site:myshopify.com", "athletic gear shopify store"],
    },
    "Surfering or Swimming Website": {
        "wp": ["surf swim sports wordpress website inurl:wp-content", "surfing blog wordpress site"],
        "sh": ["surf swim gear site:myshopify.com", "surfing equipment shopify"],
    },
    "Technologies Website": {
        "wp": ["tech blog wordpress website inurl:wp-content", "technology startup wordpress site"],
        "sh": ["tech gadgets store site:myshopify.com", "technology products shopify"],
    },
    "Transportation Website": {
        "wp": ["transport logistics wordpress website inurl:wp-content", "transportation company wordpress site"],
        "sh": ["transport accessories site:myshopify.com", "shipping transport shopify"],
    },
}

# ── Search ────────────────────────────────────────────────────────────────────
def ddg_search(query, num=15):
    links = []
    try:
        url  = f"https://html.duckduckgo.com/html/?q={urllib.parse.quote_plus(query)}"
        resp = requests.get(url, headers=HEADERS, timeout=12)
        # Pattern 1
        found = re.findall(r'href="(https?://[^"]+)"', resp.text)
        for lnk in found:
            lnk = lnk.strip()
            if not any(s in lnk for s in SKIP) and lnk not in links:
                links.append(lnk)
        if not links:
            # Pattern 2 — DDG redirect links
            found2 = re.findall(r'uddg=(https?[^&"]+)', resp.text)
            for lnk in found2:
                lnk = urllib.parse.unquote(lnk).strip()
                if not any(s in lnk for s in SKIP) and lnk not in links:
                    links.append(lnk)
    except Exception as e:
        print(f"      ⚠ DDG: {e}")
    return links[:num]

def bing_search(query, num=15):
    links = []
    try:
        url  = f"https://www.bing.com/search?q={urllib.parse.quote_plus(query)}&count={num}"
        resp = requests.get(url, headers=HEADERS, timeout=12)
        found = re.findall(r'<a[^>]+href="(https?://[^"&]+)"', resp.text)
        for lnk in found:
            lnk = lnk.strip()
            if not any(s in lnk for s in SKIP) and lnk not in links:
                links.append(lnk)
    except Exception as e:
        print(f"      ⚠ Bing: {e}")
    return links[:num]

def search(query):
    results = ddg_search(query)
    if not results:
        time.sleep(1)
        results = bing_search(query)
    return results

# ── Verify ────────────────────────────────────────────────────────────────────
def is_wordpress(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=VERIFY_TIMEOUT, allow_redirects=True)
        h = r.text.lower()
        return any(s in h for s in ["wp-content","wp-includes","/wp-json/","wordpress","wp-login"])
    except: return False

def is_shopify(url):
    try:
        r = requests.get(url, headers=HEADERS, timeout=VERIFY_TIMEOUT, allow_redirects=True)
        h = r.text.lower(); u = r.url.lower()
        return any(s in h or s in u for s in ["shopify","myshopify.com","cdn.shopify","shopify-analytics"])
    except: return False

# ── Scrape one category ───────────────────────────────────────────────────────
def scrape_cat(cat, queries, platform, seen_global):
    verify_fn = is_wordpress if platform == "wordpress" else is_shopify
    found = []

    for query in queries:
        if len(found) >= LINKS_PER_CAT: break
        print(f"      🔍 {query[:60]}")
        links = search(query)
        time.sleep(SEARCH_DELAY)

        for url in links:
            if len(found) >= LINKS_PER_CAT: break
            if url in seen_global: continue
            # Skip same domain already found
            dom = re.sub(r'https?://(www\.)?','',url).split('/')[0]
            if any(dom in u for u in found): continue

            ok = verify_fn(url)
            if ok:
                seen_global.add(url)
                found.append(url)
                plat_icon = "🔵" if platform == "wordpress" else "🟢"
                print(f"      {plat_icon} ✅ {url[:70]}")
            time.sleep(VERIFY_DELAY)

    return found[:LINKS_PER_CAT]

# ── Save Excel ────────────────────────────────────────────────────────────────
def save_excel(all_results, filename="WP_Shopify_All_Categories.xlsx"):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "WP & Shopify Sites"

    H   = PatternFill("solid", fgColor="1F4E79")
    H2  = PatternFill("solid", fgColor="2E75B6")
    WP  = PatternFill("solid", fgColor="E3F2FD")
    SH  = PatternFill("solid", fgColor="E8F5E9")
    SP  = PatternFill("solid", fgColor="D6E4F0")
    t   = Side(style="thin", color="CCCCCC")
    BRD = Border(left=t, right=t, top=t, bottom=t)
    WH  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BLD = Font(name="Arial", bold=True, size=10)
    NRM = Font(name="Arial", size=10)
    LNK = Font(name="Arial", size=10, color="0563C1", underline="single")
    WPF = Font(name="Arial", size=10, bold=True, color="21759B")
    SHF = Font(name="Arial", size=10, bold=True, color="5a8f2e")
    CTR = Alignment(horizontal="center", vertical="center")

    total = sum(len(v["wordpress"]) + len(v["shopify"]) for v in all_results.values())
    ws.merge_cells("A1:C1")
    ws["A1"] = f"WordPress & Shopify Website Links  |  {len(all_results)} Categories  |  {total} Verified Links"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = H; ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 28

    for i, h in enumerate(["Category", "Platform", "Website URL"], 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = WH; c.fill = H2; c.alignment = CTR; c.border = BRD

    row = 3
    for cat, v in all_results.items():
        wp_links = v.get("wordpress", [])
        sh_links = v.get("shopify",   [])
        if not wp_links and not sh_links: continue
        fr = row

        for url in wp_links:
            ws.cell(row=row, column=1, value=cat).fill = WP; ws.cell(row=row, column=1).border = BRD
            c2 = ws.cell(row=row, column=2, value="WordPress"); c2.fill = WP; c2.font = WPF; c2.alignment = CTR; c2.border = BRD
            c3 = ws.cell(row=row, column=3, value=url); c3.fill = WP; c3.font = LNK; c3.hyperlink = url; c3.border = BRD
            ws.row_dimensions[row].height = 18; row += 1

        for url in sh_links:
            ws.cell(row=row, column=1, value=cat).fill = SH; ws.cell(row=row, column=1).border = BRD
            c2 = ws.cell(row=row, column=2, value="Shopify"); c2.fill = SH; c2.font = SHF; c2.alignment = CTR; c2.border = BRD
            c3 = ws.cell(row=row, column=3, value=url); c3.fill = SH; c3.font = LNK; c3.hyperlink = url; c3.border = BRD
            ws.row_dimensions[row].height = 18; row += 1

        if row - 1 > fr:
            ws.merge_cells(f"A{fr}:A{row-1}")
            ws.cell(row=fr, column=1, value=cat).font = BLD
            ws.cell(row=fr, column=1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=fr, column=1).border = BRD

        for col in range(1, 4):
            ws.cell(row=row, column=col).fill = SP; ws.cell(row=row, column=col).border = BRD
        row += 1

    for col, w in zip("ABC", [32, 14, 90]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    wb.save(filename)
    print(f"\n✅ Excel saved: {filename}")

# ── Save Word Doc ─────────────────────────────────────────────────────────────
def save_doc(all_results, filename="WP_Shopify_All_Categories.docx"):
    doc_data = [{"cat": cat, "wordpress": v["wordpress"], "shopify": v["shopify"]}
                for cat, v in all_results.items()]
    total = sum(len(d["wordpress"]) + len(d["shopify"]) for d in doc_data)

    with open("_tmp_all_cats.json", "w") as f:
        json.dump({"data": doc_data, "total": total, "filename": filename}, f)

    js = r"""
const { Document, Packer, Paragraph, TextRun, ExternalHyperlink,
        HeadingLevel, AlignmentType, PageBreak } = require('docx');
const fs = require('fs');
const { data, total, filename } = JSON.parse(fs.readFileSync('_tmp_all_cats.json'));
const children = [];

children.push(
  new Paragraph({ children: [new TextRun({ text: "WordPress & Shopify Website Links", bold: true, size: 48, font: "Arial", color: "1F4E79" })], alignment: AlignmentType.CENTER, spacing: { before: 3600, after: 400 } }),
  new Paragraph({ children: [new TextRun({ text: `${data.length} Categories  |  ${total} Scraped & Verified Websites`, size: 24, font: "Arial", color: "555555" })], alignment: AlignmentType.CENTER, spacing: { after: 200 } }),
  new Paragraph({ children: [new TextRun({ text: "5 WordPress + 5 Shopify per Category", size: 20, font: "Arial", color: "888888" })], alignment: AlignmentType.CENTER, spacing: { after: 3600 } }),
  new Paragraph({ children: [new PageBreak()] })
);

data.forEach((item, idx) => {
  const { cat, wordpress: wp, shopify: sh } = item;
  if (wp.length === 0 && sh.length === 0) return;
  children.push(new Paragraph({ text: cat, heading: HeadingLevel.HEADING_2, spacing: { before: idx === 0 ? 0 : 360, after: 160 } }));
  if (wp.length > 0) {
    children.push(new Paragraph({ children: [new TextRun({ text: "WordPress Links:", bold: true, size: 22, font: "Arial", color: "21759B" })], spacing: { before: 120, after: 80 } }));
    wp.forEach(url => children.push(new Paragraph({ children: [new ExternalHyperlink({ link: url, children: [new TextRun({ text: url, style: "Hyperlink", size: 20, font: "Arial" })] })], spacing: { before: 40, after: 60 } })));
  }
  if (sh.length > 0) {
    children.push(new Paragraph({ children: [new TextRun({ text: "Shopify Links:", bold: true, size: 22, font: "Arial", color: "5a8f2e" })], spacing: { before: 160, after: 80 } }));
    sh.forEach(url => children.push(new Paragraph({ children: [new ExternalHyperlink({ link: url, children: [new TextRun({ text: url, style: "Hyperlink", size: 20, font: "Arial" })] })], spacing: { before: 40, after: 60 } })));
  }
  if (idx < data.length - 1) children.push(new Paragraph({ children: [new PageBreak()] }));
});

const doc = new Document({
  styles: {
    default: { document: { run: { font: "Arial", size: 20 } } },
    paragraphStyles: [{ id: "Heading2", name: "Heading 2", basedOn: "Normal", next: "Normal", quickFormat: true,
      run: { size: 30, bold: true, font: "Arial", color: "1F4E79" },
      paragraph: { spacing: { before: 360, after: 160 }, border: { bottom: { style: "single", size: 4, color: "2E75B6", space: 1 } }, outlineLevel: 1 }
    }]
  },
  sections: [{ properties: { page: { size: { width: 12240, height: 15840 }, margin: { top: 1080, right: 1080, bottom: 1080, left: 1080 } } }, children }]
});
Packer.toBuffer(doc).then(buf => {
  fs.writeFileSync(filename, buf);
  console.log('Doc saved: ' + filename + '  (' + total + ' links)');
});
"""
    with open("_tmp_all_cats.js", "w") as f:
        f.write(js)
    result = subprocess.run(["node", "_tmp_all_cats.js"], capture_output=True, text=True)
    print(result.stdout.strip())
    if result.returncode != 0:
        print("JS Error:", result.stderr[:300])
    for f in ["_tmp_all_cats.json", "_tmp_all_cats.js"]:
        try: os.remove(f)
        except: pass

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 65)
    print("  WORDPRESS & SHOPIFY WEBSITE SCRAPER — ALL CATEGORIES")
    print(f"  Search: DuckDuckGo + Bing")
    print(f"  Target: {LINKS_PER_CAT} WordPress + {LINKS_PER_CAT} Shopify per category")
    print(f"  Total categories: {len(CATEGORIES)}")
    print(f"  Auto-saves progress every 5 categories")
    print("=" * 65)

    # Load existing progress if any
    all_results = {}
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE) as f:
            all_results = json.load(f)
        print(f"\n  ♻️  Resuming from progress — {len(all_results)} categories done\n")

    seen_global = set()
    for v in all_results.values():
        seen_global.update(v.get("wordpress", []))
        seen_global.update(v.get("shopify",   []))

    total_wp = total_sh = 0
    cats = list(CATEGORIES.keys())

    for i, cat in enumerate(cats, 1):
        if cat in all_results:
            print(f"  [SKIP {i}/{len(cats)}] {cat} (already done)")
            continue

        print(f"\n[{i}/{len(cats)}] ══ {cat} ══")
        queries = CATEGORIES[cat]

        print(f"  🔵 WordPress:")
        wp = scrape_cat(cat, queries["wp"], "wordpress", seen_global)
        print(f"  → {len(wp)} found")

        print(f"  🟢 Shopify:")
        sh = scrape_cat(cat, queries["sh"], "shopify", seen_global)
        print(f"  → {len(sh)} found")

        all_results[cat] = {"wordpress": wp, "shopify": sh}
        total_wp += len(wp)
        total_sh += len(sh)

        # Save progress every 5 categories
        if i % 5 == 0:
            with open(PROGRESS_FILE, "w") as f:
                json.dump(all_results, f)
            print(f"\n  💾 Progress saved ({i}/{len(cats)} done)\n")

        time.sleep(1)

    # Final save
    with open(PROGRESS_FILE, "w") as f:
        json.dump(all_results, f)

    print("\n" + "=" * 65)
    total_wp = sum(len(v["wordpress"]) for v in all_results.values())
    total_sh = sum(len(v["shopify"])   for v in all_results.values())
    print(f"  TOTAL → WordPress: {total_wp}  |  Shopify: {total_sh}")
    print("=" * 65)

    save_excel(all_results)
    save_doc(all_results)

    try: os.remove(PROGRESS_FILE)
    except: pass

    print("\n🎉 Done!")
    print("   WP_Shopify_All_Categories.xlsx")
    print("   WP_Shopify_All_Categories.docx")

if __name__ == "__main__":
    main()