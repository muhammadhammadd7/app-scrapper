"""
App Link Verifier — Run this on your Mac
Reads: App_Links_Scraped.xlsx
Saves: App_Links_VERIFIED.xlsx (only valid links)

SETUP (one time):
    python3 -m pip install openpyxl pandas

RUN:
    python3 verify_links_local.py
"""

import urllib.request
import urllib.error
import time
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "App_Links_Scraped.xlsx"
OUTPUT_FILE = "App_Links_VERIFIED.xlsx"
TIMEOUT     = 10
DELAY       = 0.5   # seconds between requests (be polite to servers)

# ── URL checker ───────────────────────────────────────────────────────────────
def check_url(url: str) -> bool:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) "
            "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
        ),
        "Accept-Language": "en-US,en;q=0.9",
    }
    try:
        req = urllib.request.Request(url.strip(), headers=headers, method="HEAD")
        with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
            return resp.status in (200, 301, 302, 303)
    except urllib.error.HTTPError as e:
        # 405 = HEAD not allowed but page exists; 403 can also mean page exists
        return e.code in (301, 302, 303, 405, 403)
    except Exception:
        return False

# ── Load Excel ────────────────────────────────────────────────────────────────
def load_links(filepath):
    df = pd.read_excel(filepath, header=1)
    df['Category'] = df['Category'].ffill()
    rows = []
    for _, row in df.iterrows():
        link = str(row.get('Store Link', '')).strip()
        if 'http' not in link:
            continue
        rows.append({
            'category': str(row.get('Category', '')).strip(),
            'name':     str(row.get('App Name', '')).strip(),
            'platform': str(row.get('Platform', '')).strip(),
            'installs': str(row.get('Installs', '')).strip(),
            'rating':   str(row.get('Rating', '')).strip(),
            'link':     link,
        })
    return rows

# ── Save verified Excel ───────────────────────────────────────────────────────
def save_excel(valid_rows, total_checked, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verified Links"

    # Styles
    H    = PatternFill("solid", fgColor="1F4E79")
    H2   = PatternFill("solid", fgColor="2E75B6")
    PS   = PatternFill("solid", fgColor="E8F5E9")
    AS   = PatternFill("solid", fgColor="FFF3E0")
    t    = Side(style="thin", color="CCCCCC")
    BRD  = Border(left=t, right=t, top=t, bottom=t)
    WH   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    NRM  = Font(name="Arial", size=10)
    LNK  = Font(name="Arial", size=10, color="0563C1", underline="single")
    PSF  = Font(name="Arial", size=10, bold=True, color="1B7B34")
    ASF  = Font(name="Arial", size=10, bold=True, color="C55A11")
    CTR  = Alignment(horizontal="center", vertical="center")
    WRAP = Alignment(vertical="center", wrap_text=True)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"] = (
        f"✅ VERIFIED App Links  |  "
        f"{len(valid_rows)} valid  /  {total_checked} checked"
    )
    ws["A1"].font      = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = H
    ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 28

    # Header
    headers = ["Category", "App Name", "Platform", "Installs", "Rating", "Store Link"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = WH; c.fill = H2; c.alignment = CTR; c.border = BRD
    ws.row_dimensions[2].height = 20

    # Group by category for merged cells
    from itertools import groupby
    row_num = 3
    for cat, group in groupby(valid_rows, key=lambda x: x['category']):
        group = list(group)
        cat_start = row_num
        for app in group:
            is_ps = "Play" in app['platform']
            fill  = PS if is_ps else AS
            pf    = PSF if is_ps else ASF

            data = [
                (1, cat,            Font(name="Arial", bold=True, size=10), fill, WRAP),
                (2, app['name'],    NRM,  fill, WRAP),
                (3, app['platform'],pf,   fill, CTR),
                (4, app['installs'],NRM,  fill, CTR),
                (5, app['rating'],  NRM,  fill, CTR),
                (6, app['link'],    LNK,  fill, Alignment(vertical="center")),
            ]
            for col, val, fnt, fl, align in data:
                c = ws.cell(row=row_num, column=col, value=val)
                c.font = fnt; c.fill = fl; c.border = BRD; c.alignment = align
            # Hyperlink
            ws.cell(row=row_num, column=6).hyperlink = app['link']
            ws.cell(row=row_num, column=6).value     = app['link']
            ws.row_dimensions[row_num].height = 18
            row_num += 1

        # Merge category column
        if row_num - 1 > cat_start:
            ws.merge_cells(f"A{cat_start}:A{row_num-1}")
            ws.cell(row=cat_start, column=1).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

    # Column widths
    for col, w in zip("ABCDEF", [26, 40, 15, 16, 10, 78]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"

    wb.save(filename)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 65)
    print("  APP LINK VERIFIER")
    print(f"  Input : {INPUT_FILE}")
    print(f"  Output: {OUTPUT_FILE}")
    print("=" * 65)

    try:
        all_rows = load_links(INPUT_FILE)
    except FileNotFoundError:
        print(f"\n❌ File not found: {INPUT_FILE}")
        print("   Make sure the script is in the same folder as the Excel file.")
        return

    total = len(all_rows)
    print(f"\n  Total links to verify: {total}")
    print(f"  Play Store : {sum(1 for r in all_rows if 'Play' in r['platform'])}")
    print(f"  App Store  : {sum(1 for r in all_rows if 'App Store' in r['platform'])}")
    print("\n  Starting verification...\n")

    valid   = []
    invalid = []

    for i, row in enumerate(all_rows, 1):
        is_valid = check_url(row['link'])
        icon     = "✅" if is_valid else "❌"
        platform = "PS" if "Play" in row['platform'] else "iOS"
        print(f"  [{i:>3}/{total}] {icon} {platform} | {row['name'][:45]}")

        if is_valid:
            valid.append(row)
        else:
            invalid.append(row)

        time.sleep(DELAY)

    # Summary
    print("\n" + "=" * 65)
    print(f"  ✅ Valid   : {len(valid)}")
    print(f"  ❌ Invalid : {len(invalid)}")
    print(f"  Total     : {total}")
    print("=" * 65)

    if invalid:
        print("\n  Invalid links:")
        for r in invalid:
            print(f"    ❌ {r['platform']} | {r['name'][:40]} | {r['link'][:55]}")

    if valid:
        save_excel(valid, total, OUTPUT_FILE)
        print(f"\n🎉 Saved: {OUTPUT_FILE}  ({len(valid)} verified apps)")
    else:
        print("\n⚠️  No valid links found — check your internet connection.")

if __name__ == "__main__":
    main()