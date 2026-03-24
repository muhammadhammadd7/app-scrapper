"""
╔══════════════════════════════════════════════════════════════════╗
║         APP STORE & PLAY STORE AUTOMATED SCRAPER                ║
║  • Play Store  → google-play-scraper library                    ║
║  • App Store   → iTunes Search API (no extra library needed)    ║
║  • Filters: min 1,000 — max 500,000 downloads                   ║
║  • International links (works in USA + globally)                ║
║  • Covers 50+ categories                                        ║
║  • Saves clean Excel with clickable links                       ║
╠══════════════════════════════════════════════════════════════════╣
║  SETUP (run once):                                              ║
║    python3 -m pip install google-play-scraper openpyxl          ║
║                                                                 ║
║  RUN:                                                           ║
║    python3 app_scraper.py                                       ║
╚══════════════════════════════════════════════════════════════════╝
"""

import time, json, urllib.request, urllib.parse, warnings
warnings.filterwarnings("ignore")

try:
    from google_play_scraper import search as gp_search, app as gp_app
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError as e:
    print(f"\n❌ Missing package: {e}")
    print("Run: python3 -m pip install google-play-scraper openpyxl")
    exit(1)

MIN_INSTALLS    = 1_000
MAX_INSTALLS    = 500_000
RESULTS_PER_CAT = 20
DELAY           = 1.2

# Multi-country search for international coverage
# Play Store: search in multiple countries, deduplicate by package ID
# App Store: use 'us' country but results work globally
PLAY_COUNTRIES  = ["us", "gb", "in"]   # USA, UK, India — covers global apps
IOS_COUNTRY     = "us"                  # iTunes US results are globally accessible
LANGUAGE        = "en"

CATEGORIES = {
    "Accounting or Book Keeping": [
        "small business accounting invoice maker",
        "bookkeeping expense tracker freelancer",
        "invoice billing accounting app",
        "accounting software small business",
    ],
    "Agriculture or Farming": [
        "farming crop management app",
        "agriculture field tracker farmer",
        "farm management livestock tracker",
        "precision agriculture app farmer",
    ],
    "Art or Sketch": [
        "sketch drawing art app",
        "digital painting illustration app",
        "comic manga drawing app",
        "art canvas drawing creative app",
    ],
    "Baby": [
        "baby tracker feeding sleep log",
        "newborn care milestone tracker",
        "baby monitor breastfeeding app",
        "baby development growth tracker",
    ],
    "Bakery and Cake": [
        "bakery order management cake business",
        "cake decorating recipe baker app",
        "bakery pos point of sale system",
        "cake shop order management app",
    ],
    "Bakery": [
        "bakery delivery ordering app",
        "bread shop management app",
        "local bakery food ordering",
        "pastry bakery business management",
    ],
    "Bicycle": [
        "cycling route tracker bike gps",
        "mountain bike trail navigation",
        "bicycle repair maintenance log",
        "bike riding tracker cycling app",
    ],
    "Blogging": [
        "blogging writing platform app",
        "blog post creator writer app",
        "content writing blogging tool",
        "blog creator publish articles app",
    ],
    "Business or Investment": [
        "small business investment tracker",
        "stock portfolio investment app",
        "business finance management app",
        "investment portfolio tracker app",
    ],
    "Candy": [
        "candy sweets shop ordering app",
        "confectionery pos retail app",
        "sweet shop loyalty rewards app",
        "candy store management app",
    ],
    "Car or Automobile": [
        "car maintenance log mileage tracker",
        "used car buying selling app",
        "auto repair shop management",
        "vehicle car service tracker app",
    ],
    "Care": [
        "caregiver elderly care management",
        "home care scheduling app",
        "senior care family app",
        "elderly patient care tracker app",
    ],
    "CBD Oil": [
        "cbd dosage tracker wellness app",
        "hemp cbd journal tracker",
        "cannabis wellness dosage app",
        "cbd oil health tracker app",
    ],
    "Charity Foundation": [
        "charity donation fundraising app",
        "nonprofit volunteer management",
        "crowdfunding charity giving app",
        "donation fundraiser nonprofit app",
    ],
    "Cleaning": [
        "cleaning service scheduling app",
        "house cleaning booking management",
        "maid service field management",
        "home cleaning service app",
    ],
    "Clothing": [
        "wardrobe outfit planner app",
        "clothing resale secondhand marketplace",
        "fashion style closet organizer",
        "clothing store management app",
    ],
    "Coaching": [
        "life coach habit tracking app",
        "sports coaching training app",
        "personal coach session management",
        "coaching client management app",
    ],
    "Consultant": [
        "consultant client management app",
        "freelance consulting invoice app",
        "business consulting crm app",
        "consulting project management app",
    ],
    "Cosmetic or Beauty": [
        "makeup beauty product scanner app",
        "skincare cosmetic ingredient checker",
        "virtual makeup beauty try on",
        "beauty cosmetics shopping app",
    ],
    "Costume or Swimwear": [
        "outfit costume ideas fashion app",
        "swimwear fashion style app",
        "dress up costume planner",
        "fashion costume design app",
    ],
    "Dating": [
        "dating app meet people nearby",
        "niche dating relationship app",
        "local singles dating app",
        "dating match relationship app",
    ],
    "Digital or Marketing": [
        "social media scheduler marketing app",
        "email marketing small business app",
        "digital marketing analytics tool",
        "social media marketing management app",
    ],
    "Doctor or Clinic": [
        "doctor appointment booking clinic app",
        "telemedicine clinic patient app",
        "medical appointment scheduler",
        "online doctor consultation app",
    ],
    "Dog Websites": [
        "dog care training tracker app",
        "pet dog health record app",
        "dog walking booking app",
        "dog training pet care app",
    ],
    "Ecommerce": [
        "ecommerce shop online selling app",
        "marketplace buy sell local app",
        "online store management seller",
        "ecommerce store builder app",
    ],
    "Education": [
        "online learning courses education app",
        "skill learning course platform app",
        "tutoring education learning app",
        "school learning student education app",
    ],
    "Engineering": [
        "engineering calculator tools app",
        "field service engineering app",
        "technical calculator engineering",
        "civil engineering tools app",
    ],
    "Event Planning": [
        "event planning management app",
        "wedding event organizer app",
        "venue booking event scheduler",
        "event coordinator management app",
    ],
    "Fitness Gym": [
        "gym workout tracker fitness app",
        "personal training workout planner",
        "gym management membership app",
        "fitness workout exercise tracker app",
    ],
    "Food Restaurant": [
        "restaurant food ordering app",
        "local restaurant booking reservation",
        "food delivery small restaurant",
        "restaurant management pos app",
    ],
    "Furniture": [
        "furniture interior design ar app",
        "home decor furniture shopping app",
        "room planner interior design",
        "furniture store shopping app",
    ],
    "Gaming": [
        "indie puzzle game mobile app",
        "casual strategy game app",
        "adventure indie mobile game",
        "mobile casual puzzle game app",
    ],
    "Health or Nutrition": [
        "health nutrition tracker app",
        "symptom tracker medication reminder",
        "calorie nutrition food tracker",
        "health wellness tracker app",
    ],
    "Hotel or Traveling": [
        "travel planner trip organizer app",
        "hotel booking travel deals app",
        "budget travel trip tracker",
        "travel booking vacation app",
    ],
    "Jewellery": [
        "jewelry shopping app online store",
        "jewellery design catalog app",
        "custom jewelry order app",
        "jewelry store management app",
    ],
    "Logistics": [
        "delivery route optimization app",
        "logistics driver delivery management",
        "courier tracking dispatch app",
        "freight logistics tracking app",
        "supply chain logistics management",
        "warehouse inventory logistics app",
    ],
    "Music or Event": [
        "indie music streaming app",
        "concert ticket event booking app",
        "music discovery artist app",
        "music event ticket booking app",
    ],
    "Perfume": [
        "perfume fragrance community app",
        "perfume discovery subscription app",
        "fragrance review tracker app",
        "perfume store shopping app",
    ],
    "Photography": [
        "photo editing camera app",
        "photography client gallery app",
        "photo filter editing app",
        "camera photo studio editing app",
    ],
    "Plumbing or Heating": [
        "plumber field service management app",
        "hvac technician job management",
        "plumbing service booking app",
        "plumber heating service app",
    ],
    "Real Estate or Construction": [
        "construction project management app",
        "real estate property listing app",
        "contractor site management app",
        "property real estate buying app",
    ],
    "Recruitment or Job Finding": [
        "job search hiring app",
        "freelance gig work finding app",
        "part time job recruitment app",
        "job portal recruitment app",
    ],
    "Saloon Spa": [
        "salon appointment booking app",
        "spa beauty booking management",
        "barber appointment scheduling app",
        "hair salon booking app",
    ],
    "Skin or Hair": [
        "skincare routine tracker app",
        "hair care routine beauty app",
        "skin analysis beauty app",
        "skin hair care tracker app",
    ],
    "Solar Panel": [
        "solar energy monitoring app",
        "solar panel production tracker",
        "solar installation management app",
        "solar power energy tracker app",
    ],
    "Sports": [
        "sports team management app",
        "sports score tracking app",
        "amateur sports league app",
        "sports coaching management app",
    ],
    "Surfing or Swimming": [
        "surf forecast wave tracker app",
        "swimming workout tracker app",
        "water sports surf community app",
        "swim training tracker app",
    ],
    "Technologies": [
        "developer tools coding app",
        "tech news product discovery app",
        "programming learning code app",
        "software developer productivity app",
    ],
    # ── TRANSPORTATION — expanded: taxi + logistics + freight + delivery ──
    "Transportation": [
        "logistics freight shipping tracking app",
        "truck driver logistics management app",
        "fleet management vehicle tracking app",
        "cargo shipping logistics app",
        "delivery route driver app",
        "ride booking taxi local app",
        "bus transport public transit app",
        "moving relocation transport app",
    ],
    "ALL WEBSITE": [
        "website builder creator app",
        "drag drop website builder app",
        "website design template app",
        "no code website builder app",
    ],
}

# ─── Helpers ──────────────────────────────────────────────────────────────────
def parse_installs(val):
    if not val: return 0
    s = str(val).replace(",","").replace("+","").strip().upper()
    try:
        if "B" in s: return int(float(s.replace("B",""))*1_000_000_000)
        if "M" in s: return int(float(s.replace("M",""))*1_000_000)
        if "K" in s: return int(float(s.replace("K",""))*1_000)
        return int(s)
    except: return 0

def in_range(n): return MIN_INSTALLS <= n <= MAX_INSTALLS
def gplay_link(a): return f"https://play.google.com/store/apps/details?id={a}"
def appstore_link(a): return f"https://apps.apple.com/app/id{a}"  # No country code = works globally

# ─── Play Store scraper (multi-country) ───────────────────────────────────────
def scrape_playstore(cat_name, queries):
    print(f"  ▶ Play Store — {cat_name}")
    seen, valid = set(), []
    for country in PLAY_COUNTRIES:
        if len(valid) >= 7: break
        for q in queries:
            if len(valid) >= 7: break
            try:
                results = gp_search(q, lang=LANGUAGE, country=country, n_hits=RESULTS_PER_CAT)
            except Exception as e:
                print(f"    ⚠ [{country}] {e}")
                time.sleep(2)
                continue
            for r in results:
                if len(valid) >= 7: break
                aid = r.get("appId", "")
                if not aid or aid in seen: continue
                seen.add(aid)
                try:
                    d = gp_app(aid, lang=LANGUAGE, country=country)
                    ri = d.get("realInstalls", 0) or parse_installs(d.get("installs", "0"))
                    title = d.get("title", r.get("title", ""))
                    score = round(d.get("score", 0) or 0, 1)
                    istr  = d.get("installs", "N/A")
                except:
                    ri    = parse_installs(r.get("installs", "0"))
                    title = r.get("title", "")
                    score = round(r.get("score", 0) or 0, 1)
                    istr  = r.get("installs", "N/A")
                if in_range(ri):
                    valid.append({
                        "name": title,
                        "installs_str": istr,
                        "score": score,
                        "link": gplay_link(aid)
                    })
                    print(f"    ✅ [{country}] {title[:38]:38} | {istr:12} | ⭐{score}")
                time.sleep(DELAY)
    return valid[:7]

# ─── App Store scraper (international link format) ─────────────────────────────
def scrape_appstore(cat_name, queries):
    print(f"  🍎 App Store  — {cat_name}")
    seen, valid = set(), []
    for q in queries:
        if len(valid) >= 7: break
        try:
            url = (
                f"https://itunes.apple.com/search"
                f"?term={urllib.parse.quote(q)}"
                f"&entity=software&limit={RESULTS_PER_CAT}&country={IOS_COUNTRY}"
            )
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                results = json.loads(resp.read().decode()).get("results", [])
        except Exception as e:
            print(f"    ⚠ {e}")
            time.sleep(2)
            continue
        for r in results:
            if len(valid) >= 7: break
            aid = str(r.get("trackId", ""))
            if not aid or aid in seen: continue
            seen.add(aid)
            title = r.get("trackName", "")
            rc    = r.get("userRatingCount", 0) or 0
            score = round(r.get("averageUserRating", 0) or 0, 1)
            est   = rc * 70
            if in_range(est):
                # Use /app/id format (no country) — works internationally
                valid.append({
                    "name": title,
                    "installs_str": f"~{est:,}",
                    "score": score,
                    "link": appstore_link(aid)
                })
                print(f"    ✅ {title[:40]:40} | ~{est:>9,} | ⭐{score}")
        time.sleep(DELAY)
    return valid[:7]

# ─── Save Excel ───────────────────────────────────────────────────────────────
def save_excel(all_results, filename="App_Links_Scraped.xlsx"):
    wb = Workbook(); ws = wb.active; ws.title = "App Links"
    H  = PatternFill("solid", fgColor="1F4E79")
    H2 = PatternFill("solid", fgColor="2E75B6")
    PS = PatternFill("solid", fgColor="E8F5E9")
    AS = PatternFill("solid", fgColor="FFF3E0")
    SP = PatternFill("solid", fgColor="D6E4F0")
    t  = Side(style="thin", color="CCCCCC")
    BRD = Border(left=t, right=t, top=t, bottom=t)
    WH  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    BLD = Font(name="Arial", bold=True, size=10)
    NRM = Font(name="Arial", size=10)
    LNK = Font(name="Arial", size=10, color="0563C1", underline="single")
    PSF = Font(name="Arial", size=10, bold=True, color="1B7B34")
    ASF = Font(name="Arial", size=10, bold=True, color="C55A11")
    CTR = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Auto-Scraped Apps  |  {MIN_INSTALLS:,} - {MAX_INSTALLS:,} downloads  |  International Links  |  50+ Categories"
    ws["A1"].font = Font(name="Arial", bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = H
    ws["A1"].alignment = CTR
    ws.row_dimensions[1].height = 28

    for i, h in enumerate(["Category", "App Name", "Platform", "Installs", "Rating", "Store Link"], 1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = WH; c.fill = H2; c.alignment = CTR; c.border = BRD

    row = 3
    for cat, data in all_results.items():
        ps_apps = data.get("playstore", [])
        as_apps = data.get("appstore", [])
        if not ps_apps and not as_apps: continue
        fr = row
        for app in ps_apps:
            for col, val, fnt, fill in [
                (2, app["name"],         NRM, PS),
                (3, "Play Store",        PSF, PS),
                (4, app["installs_str"], NRM, PS),
                (5, f"{app['score']}",   NRM, PS),
                (6, app["link"],         LNK, PS),
            ]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = fnt; c.fill = fill; c.border = BRD
                if col in (3, 4, 5): c.alignment = CTR
            ws.cell(row=row, column=6).hyperlink = app["link"]
            ws.cell(row=row, column=1).fill = PS
            ws.cell(row=row, column=1).border = BRD
            row += 1
        for app in as_apps:
            for col, val, fnt, fill in [
                (2, app["name"],         NRM, AS),
                (3, "App Store",         ASF, AS),
                (4, app["installs_str"], NRM, AS),
                (5, f"{app['score']}",   NRM, AS),
                (6, app["link"],         LNK, AS),
            ]:
                c = ws.cell(row=row, column=col, value=val)
                c.font = fnt; c.fill = fill; c.border = BRD
                if col in (3, 4, 5): c.alignment = CTR
            ws.cell(row=row, column=6).hyperlink = app["link"]
            ws.cell(row=row, column=1).fill = AS
            ws.cell(row=row, column=1).border = BRD
            row += 1
        if row - 1 >= fr:
            if row - 1 > fr:
                ws.merge_cells(f"A{fr}:A{row-1}")
            ws.cell(row=fr, column=1, value=cat).font = BLD
            ws.cell(row=fr, column=1).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            ws.cell(row=fr, column=1).border = BRD
        for col in range(1, 7):
            ws.cell(row=row, column=col).fill = SP
            ws.cell(row=row, column=col).border = BRD
        row += 1

    for col, w in zip("ABCDEF", [26, 38, 14, 18, 10, 75]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    wb.save(filename)
    print(f"\n✅ Saved: {filename}")


# ─── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("=" * 65)
    print(f"  APP SCRAPER  |  {MIN_INSTALLS:,} – {MAX_INSTALLS:,} downloads")
    print(f"  Countries (Play Store): {', '.join(PLAY_COUNTRIES)}")
    print(f"  App Store: International links (no country code)")
    print(f"  Categories: {len(CATEGORIES)}")
    print("=" * 65)

    all_results = {}
    for i, (cat, queries) in enumerate(CATEGORIES.items(), 1):
        print(f"\n[{i}/{len(CATEGORIES)}] ══ {cat} ══")
        ps = scrape_playstore(cat, queries)
        ap = scrape_appstore(cat, queries)
        all_results[cat] = {"playstore": ps, "appstore": ap}
        print(f"  → PS:{len(ps)} | AS:{len(ap)}")
        time.sleep(1)

    total_ps = sum(len(v["playstore"]) for v in all_results.values())
    total_as = sum(len(v["appstore"])  for v in all_results.values())
    print("\n" + "=" * 65)
    print(f"  TOTAL → Play Store: {total_ps}  |  App Store: {total_as}")
    print("=" * 65)
    save_excel(all_results)
    print("\nDone! Open App_Links_Scraped.xlsx")

if __name__ == "__main__":
    main()